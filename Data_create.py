import os
import openpyxl
from tika import parser
from openpyxl import Workbook, load_workbook

keywords = ['משחק', 'שחקן' , 'שחקנים' , 'קרב',  'ניצחון',  'הפסד',  'Unity', 'גיימינג' , 'עלילה', 'פאזל', 'פאזלים' , 'עלילתי' ,
'חייזר' , 'שחקני' ,'שלב' , 'יוניטי', 'מסתורי', 'ניקוד', 'אסטרטגיה', 'יריב',  'תחרות', 'משימה',  'שלב', 'אויב' 'הרפתקאות',
'הרפתקה', 'מכשולים',  'להילחם', 'פסילה', 'חייל','משפטיים' , 'משפטיות' , 'משפטי', 'משפטית', 'עורך' , 'עורכי' , 'דין' , 'משפט' ,'זכות',
'זכויות' , 'תביעה' , 'חוקים', 'חקיקה' , 'נחקק' , 'חוק' , 'סעיף' , 'פרקליט', 'עדות' , 'עדויות'  'ראיות' , 'דיון' , 'דיונים' , 'תיק' ,
'מרצה' ,'מרצים','אקדמיה', 'מחלקה' , 'ועדה' , 'הרצאות', 'מצגות', 'מצגת', 'בתי ספר' , 'בתי הספר', 'בית הספר' , 'בית ספר' , 'לימודים' , 'תלמיד' ,
'קורס', 'מקצוע','מורים' ,'מורה', 'הוראה', 'שיעור', 'חינוך', 'כיתה', 'כיתתי', 'ילד' , 'ילדים', 'זיווג', 'הורות' , 'משפחה' , 'זוג' ,'לידה' , 'שידוך',
'שידוכים', 'זוגיות', 'בן זוג', 'בת זוג', 'היכרות', 'תינוק' , 'תור'  , 'רופא' , 'מטופל' , 'מטפל' , 'טיפול', 'רפואה' , 'רפואי' , 'מדד' , 'לחץ דם' , 'דופק' ,
'סטורציה' , 'נשימה' , 'נשימות', 'אבחון' , 'הפרעה' , 'הפרעות', 'אירידיולוגיה','מוקדים' , 'טלפוניים',  'טלפון', 'מייל' , 'דיווח', 'לדווח' , 'מוקד' , 'פניה' ,
'פניות' , 'מדווח' , 'להגיב' , 'מסמכים', 'טקסטים' , 'טקסט' , 'טקסטואליות', 'טקסטואלית', 'מסמך', 'תרגום', 'ביטוי', 'חילוץ', 'ביטויים', 'ארכיון', 'ארכיונים', 'קריאה',
'כבדי שמיעה', 'כבד השמיעה' , 'לקות',  'שפת הסימנים', 'חירש' , 'בעלי מוגבלות' , 'לקויות' , 'עיוורים' , 'אוטיזם' , 'אוטיסט' , 'נויורולוגי' , 'נחייה', 'אילם', 'אילמים', 
'קהל', 'אנשים', 'שאלונים' , 'שאלון', 'סקר', 'התנדבות' , 'אוכלוסייה' , 'מתנדב' , 'חברתי' , 'תוכן', 'תכנים' , 'שיתופי',
'רכישה', 'חנות' , 'מפעל' , 'יצרן' , 'יצרנים', 'מחיר' , 'עסק' ,'מסחר' , 'משלוח' , 'חשבונית' , 'חשבוניות' , 'יד שניה', 'מכירה', 'מוצר', 'מפרסם', 'עלות', 'תשלום',
'פריט', 'שוק', 'סופרמרקט', 'קופה', 'קופאי', 'ספק', 'חנויות', 'הזמנה', 'עובד' , 'כוח אדם' , 'מסגרת' , 'הכשרה' ,'מוזיקלי', 'מוזיקה' , 'קול' , 'אפקט', 'סאונד' , 'הופעה' ,
'תייר' ,'מסורת' , 'תרבות' , 'אטרקציות' , 'אטרקציה' , 'מסעדה'  , 'מסעדות' , 'מוזיאון' , 'סרט','Movie', 'אמן' , 'אמנים' ,'כרטיס', 'ספריה', 'מוזמנים', 'אירוע', 'סאונד', 
'תופים', 'נגינה','מזון', 'ירקות', 'פירות', 'בשר', 'חלב', 'מתכונים', 'מתכון',
'בריאות', 'כושר', 'אימון', 'מתאמן', 'תזונה', 'מאמנים', 'עישון',  'מעשן', 'גמילה', 'מעשנים', 'סיגריה','כנס', 'סיגריות',
'מתמטי' , 'פונקציה' , 'פונקציות' , 'אלגברה' , 'מטריצה' , 'דטרמיננטה' , 'גרף', 'גרפים', 'קודקוד',
'דירה' , 'דייר' , 'השכרה', 'השכרת', 'דירות' , 'שוכר', 'משכיר', 'חשמל','ארנונה', 'משכנתא', 'הוצאות', 'הכנסות', 'חסכונית', 'חשבון', 'חשבונות', 'דוח',  
'עיצוב' , 'לעצב' , 'יעצב' , 'תמונה' , 'תמונות' , 'פילטר' , 'filter' , 
'נהג' , 'צמיג' , 'רכב', 'מוסך' , 'נסיעות' , 'יעד' , 'ניווט' , 'waze' , 'לנסוע' ,'צמח','כביש' , 'דרך'  , 'דרכים' , 'הגה' , 'נהיגה' , 'נתיב', 'נוסע', 'נסיעה', 'מפה', 'מצפן', 'מיקום',
'מרחק', 'מסלול', 'טיול' , 'הליכה' , 'טיול' , 'כלב' , 'סביבה',
'סכנות' ,'מסוכן', 'סכנה' , 'נזק' , 'שריפה', 'שריפות', 'להתריע' , 'התרעה', 'חירום', 'הצלה', 'לכוד', 'לכודים',  'מצוקה', 'בריונות' , 'אלימות' , 'אלימה' , 'אלים' , 'התעללות' ,
'קוד', 'שגיאה', 'מתכנת',  'stack overflow', 'אובייקט', 'תחזוקה', 'שרת', 'חומרה']


#load the excel table , fill the first row with "file name" and keywords
def create_excel_table():
    
    # Create a new Excel workbook
    data_work_book = load_workbook("data_table.xlsx")

    # Select the active worksheet
    work_sheet = data_work_book.active

    # Set the first cell of the first row to 'File Name'
    work_sheet.cell(row=1, column=1).value = 'File Name'

    # Add keywords to the first row
    for i, keyword in enumerate(keywords):
        work_sheet.cell(row=1, column=i + 2).value = keyword

    return data_work_book, work_sheet


def extract_text_from_pdf(pdf_file_path):
    parsedPDF = parser.from_file(pdf_file_path)
    return parsedPDF['content']


def create_text_file(pdf_file_path, extracted_text):
    #create the name of the new text file
    text_file_name = os.path.splitext(pdf_file_path)[0] + '.txt'

    with open(text_file_name, 'w', encoding='utf-8') as text_file:
        text_file.write(extracted_text)


def add_data_to_excel_table(worksheet, next_empty_row_number, pdf_file_name, keyword_counts_list):
    #put the PDF file name
    worksheet.cell(row=next_empty_row_number, column=1).value = pdf_file_name
    for i, count in enumerate(keyword_counts_list):
        worksheet.cell(row=next_empty_row_number, column=i + 2).value = count


def process_folder(main_folder, worksheet):

    # Find the next empty row in the worksheet - we'll put the final data there
    next_empty_row_number = worksheet.max_row + 1

    # loop through each folder in the main folder
    for sub_folder_name in os.listdir(main_folder):
        sub_folder_path = os.path.join(main_folder, sub_folder_name)
        # check if folderpath is a folder (not a file)
        if os.path.isdir(sub_folder_path):
            print('Extracting text from PDFs in folder:', sub_folder_name)
            # loop through each PDF file in the folder
            for pdf_file_name in os.listdir(sub_folder_path):
                if pdf_file_name.endswith('.pdf'):
                    pdf_file_path = os.path.join(sub_folder_path, pdf_file_name)
                    # extract text from the PDF file
                    extracted_text = extract_text_from_pdf(pdf_file_path)
                    print('Text extracted from file:', pdf_file_name)

                    # create a new text file with the same name as the PDF file
                    create_text_file(pdf_file_path, extracted_text)

                    # Count keyword occurrences in the extracted text
                    keyword_counts_list = [1 if keyword in extracted_text else 0 for keyword in keywords]
 
                    # Add the file name and keyword counts to the Excel table
                    add_data_to_excel_table(worksheet, next_empty_row_number, pdf_file_name, keyword_counts_list)
                    next_empty_row_number += 1

    return worksheet


# set the path to your main folder containing subfolders with PDF files
main_folder = 'C:/Users/lidor/Documents/Lidor/Lidor Study/Project Portal/data'

# load the Excel table
workbook, worksheet = create_excel_table()

#Process the main folder and add data to the Excel table
worksheet_with_keywords_counting = process_folder(main_folder, worksheet)

# final_table_with_topics = add_topic_to_table(worksheet_with_keywords_counting, topic)

#Save the workbook
workbook.save('data_table.xlsx')


# Count keyword occurrences in the extracted text
                    # keyword_counts_list = [extracted_text.count(keyword) for keyword in keywords]


#add the 'topic' column to the data table
# def add_topic_to_table(worksheet, topic):
#     # Create a new column titled 'Topic'
#     worksheet.cell(row=1, column=len(keywords) + 2).value = 'Topic'
    
#     # Add each topic to the new column
#     for i, topic in enumerate(topic):
#         worksheet.cell(row=i + 2, column=len(keywords) + 2).value = topic
        
#     return worksheet

# topic = ['Gaming and science fiction','documents and text','Family and relationship','Leisure and culture','criminal and emergency',
#           'Gaming and science fiction','housework','market','Software&Hardware','Gaming and science fiction','school and academia','Gaming and science fiction',
#           'social','school and academia','market','market','Health and sport','Software&Hardware','criminal and emergency','Gaming and science fiction',
#           'Software&Hardware','Support for the disabled','market','Nature , navigation and driving','Leisure and culture','school and academia'
#           ,'Support for the disabled','Family and relationship','criminal and emergency','Nature , navigation and driving','Gaming and science fiction'
#           ,'Health and sport','Nature , navigation and driving','documents and text','Software&Hardware','Gaming and science fiction' ,'Leisure and culture','communication'
#           ,'Software&Hardware','Gaming and science fiction','Leisure and culture','Software&Hardware','Design','Nature , navigation and driving','medicine'
#           ,'Gaming and science fiction','Nature , navigation and driving','market','Gaming and science fiction','Design','Leisure and culture','Gaming and science fiction'
#           ,'Gaming and science fiction','Gaming and science fiction','math','math','Support for the disabled','criminal and emergency','Nature , navigation and driving','Leisure and culture',
#           'medicine','Gaming and science fiction' ,'Nature , navigation and driving','laws','Design','Software&Hardware','Gaming and science fiction'
#           ,'Support for the disabled','Gaming and science fiction','market','medicine','school and academia','medicine' ,'criminal and emergency','social' ,'Support for the disabled'
#           ,'laws' ,'Nature , navigation and driving','social','laws','school and academia','Family and relationship','Leisure and culture','Gaming and science fiction'
#           ,'medicine'  ,'Nature , navigation and driving','communication','Leisure and culture','Gaming and science fiction','Gaming and science fiction','laws','Gaming and science fiction'
#           ,'documents and text','Gaming and science fiction' ,'Support for the disabled','social','Gaming and science fiction','market','market','Leisure and culture' ,'documents and text'
#           ,'social','school and academia','market','criminal and emergency','housework']