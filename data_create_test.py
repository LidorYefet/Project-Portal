import os
import openpyxl
from tika import parser
from openpyxl import Workbook, load_workbook

keywords = ['משחק', 'שחקן' , 'שחקנים' , 'קרב',  'ניצחון',  'הפסד',  'Unity', 'גיימינג' , 'עלילה','Game', 'פאזל', 'פאזלים' , 'עלילתי' , 'מכשפה' ,
'חייזרים' , 'חייזר' , 'קסומה' , 'קסמים'  ,'קסם', 'שחקני' ,'שלב' , 'יוניטי', 'מסתורי', 'פוטבול', 'ניקוד', 'משחקי מחשב', 'אסטרטגיה',
'יריב', 'להביס', 'אמצעי לחימה', 'תחרות', 'משימה', 'לשרוד', 'הישרדות', 'דרגה', 'דרגת קושי',  'שלב', 'אויב', 'אויבים', 'player',
'נשק', 'מגן', 'חיים', 'הרפתקאות', 'הרפתקה', 'גיבור', 'מכשולים', 'VR', 'מרובה משתתפים', 'להילחם', 'יריות', 'פסילה', 'חייל',
'כישופים', 'כישוף', 'משפטיים' , 'משפטיות' , 'משפטי', 'משפטית', 'משפטית', 'עורך' , 'עורכי' , 'דין' , 'משפט' , 'זכויות' , 'תביעה' ,
'חוקים', 'חקיקה' , 'המשפטים' , 'נחקקים' , 'חוק' , 'סעיף' , 'פרקליט', 'עדות' , 'עדויות' , 'ראיה', 'ראיות' , 'דיון' , 'דיונים' ,
'תיק' ,'מרצה' , 'מרצים', 'מחלקתו', 'אקדמיה', 'מחלקה' , 'ועדה' , 'הרצאות', 'מצגות', 'מצגת', 'בתי ספר' , 'בתי הספר', 'בית הספר' ,
'בית ספר' , 'לימודים' , 'תלמיד' , 'כיתה', 'קורס', 'מקצוע', 'מקצועות', 'מורים' ,'מורה', 'הוראה', 'שיעור', 'שיעורים און ליין',
'חינוך', 'KAHOOT', 'כיתה', 'כיתתי', 'ילד' , 'ילדים', 'זיווגים', 'הורות' , 'משפחה' , 'תינוק','זוג' ,'לידה' , 'שידוך', 'זוגיות',
'בן זוג', 'בת זוג', 'היכרות', 'אישיות', 'גברים', 'נשים', 'פרופיל אישי',  'תור' , 'מטופלים',  'מטופל' , 'מטפל' , 'טיפול',
'רפואה' , 'רפואי' , 'מדדים' , 'לחץ דם' , 'דופק' , 'סטורציה' , 'נשימה' , 'נשימות', 'אבחון' , 'הפרעה' , 'הפרעות', 'אירידיולוגיה',
'מסלול', 'טיול' , 'הליכה' , 'טיולי' , 'כלב' , 'כלבים', 'חיות שאבדו','מוקדים' , 'טלפוניים',  'טלפון', 'מייל' , 'דיווח',
'לדווח' , 'מוקד' , 'פניה' , 'פניות' , 'מדווח' , 'להגיב' , 'מסמכים', 'טקסטים' , 'טקסט' , 'טקסטואליות', 'טקסטואלית',
'מסמך', 'תרגום', 'ביטוי', 'שפה', 'חילוץ', 'ביטויים', 'ארכיון', 'ארכיונים', 'קריאה', 'קטלוג', 'כבדי שמיעה', 'כבד השמיעה' ,
'שפת הסימנים', 'חירש' , 'בעלי מוגבלות' , 'לקויות' , 'עיוורים' , 'אוטיזם' , 'אוטיסט' , 'נויורולוגי' , 'נחייה', 'אילם',
'אילמים', 'קהל', 'אנשים', 'שאלונים' , 'שאלון', 'סקר', 'שאלות', 'התנדבות' , 'אוכלוסייה' , 'מתנדב' , 'רשת חברתית' , 
'תוכן', 'תכנים' , 'שיתופי','בריונות ברשתות חברתיות', 'בריונות' , 'בריונות ברשת', 'אלימות', 'אלים', 'פוסט אלים',
'מוזיקלי', 'מוזיקה' , 'מוזיקליים', 'נגינה' , 'אפקט', 'גל קול' , 'סאונד' , 'הופעה' , 'תייר' , 'מסורת' , 'תרבות' ,
'אטרקציות' , 'אטרקציה' , 'מסעדה'  , 'מסעדות' , 'מוזיאון' , 'סרט','Movie', 'אמן' , 'אמנים' ,'כרטיס', 'ספריה',
'סופר', 'wedding', 'אישורי הגעה', 'סידורי הושבה', 'מוזמנים', 'אירוע', 'סטאטוס הגעה', 'סידורי השולחן', 'רכישה',
'חנות', 'רכישת' , 'מפעל' , 'יצרן' , 'יצרנים' , 'מוצר', 'מחיר' , 'עסק', 'לקוח' ,'מסחר' , 'משלוח' , 'חשבונית' ,
'חשבוניות' , 'יד שניה', 'מכירה', 'מוצר', 'מפרסם', 'עלות', 'תשלום', 'פריט', 'שוק', 'פירות וירקות', 'סופרמרקט', 'קופה', 'קופאי',
'ספק', 'חנויות', 'הזמנה','בריאות', 'כושר גופני', 'כושר', 'אימון', 'אימון אישי', 'מאמן', 'מתאמן', 'תזונה', 'מאמן אישי', 'מאמנים',
'חדר כושר', 'עישון',  'מעשן', 'גמילה', 'מעשנים', 'סיגריה', 'סיגריות', 'סיגריה','עובד' , 'כוח אדם' , 'מסגרת' , 'הכשרה' ,'מתמטי' ,
'פונקציה' , 'פונקציות' , 'אלגברה' , 'מטריצה' , 'דטרמיננטה' , 'גרף', 'גרפים', 'קודקוד','דירה' , 'דייר' , 'השכרה', 'השכרת', 'דירות' ,
'שוכר', 'משכיר', 'חשמל', 'מים', 'ארנונה', 'משכנתא', 'הוצאות', 'הכנסות', 'משק בית', 'חסכונית', 'חשבון', 'חשבונות', 'דוח',
'ממוצע','אלימות' , 'אלימה' , 'אלים' , 'התעללות' , 'עיצוב' , 'לעצב' , 'יעצב' , 'תמונה' , 'תמונות' , 'פילטר' , 'filter' , 'נהג' , 'צמיג' ,
'רכב', 'מוסך' ,  'נסיעות' , 'יעד' , 'ניווט' , 'waze' , 'לנסוע' ,'כביש' , 'דרך'  , 'דרכים' , 'הגה' , 'נהיגה' , 'נתיב', 'צופר', 'נוסע', 'איסוף',
'נסיעה', 'google maps', 'מפה', 'מצפן', 'שטח פתוח', 'מיקום', 'מרחק','סכנות' ,'מסוכן', 'סכנה' , 'פגיעה בנפש', 'חיי אדם', 'נזק' , 'דליפות גז',
'דליפה', 'שריפה', 'שריפות', 'להתריע' , 'התרעה', 'כוחות חירום', 'הצלה', 'לכוד', 'לכודים', 'סכנות מידיות', 'חיישנים להתרעה', 'מצוקה'
, 'אותות', 'איתור',  'קוד', 'שורות קוד', 'שגיאה', 'שגיאות זמן ריצה', 'זמן ריצה', 'זמן פיתוח', 'מתכנת', 'המתכנתים',
'פתרון', 'שגיאה לשרת', 'stack overflow', 'אובייקט', 'תחזוקה', 'שרתים', 'שרת', 'תחזוקת שרתים', 'ביצועי שרתים', 'ניטור שרתים', 
'אסוף לוגים', 'שרתים', 'ניתוחים סטטיסטים','סביבה', 'מזון נזרק', 'פגיעה', 'פגיעה בסביבה','מזון', 'טרי', 'פג תוקף', 'ירקות', 
'פירות', 'בשר', 'חלב', 'מוצרי חלב', 'צמחוני', 'מתכונים', 'מתכון','דחיסה', 'דחיסת הופמן', 'אלגוריתם דחיסה',
'מחשב', 'פריטי חומרה', 'מהירות', 'שימוש ביתי', 'גיימר', 'שדרוג', 'קנייה', 'הרכבה', 'המלצות','סאונד', 'תופים', 'מערכת תופים', 'נגינה', 'לימוד נגינה' ]  
topic = ['A' , 'B']

#load the excel table , fill the first row with "file name" and keywords
def create_excel_table():
    
    # Create a new Excel workbook
    data_work_book = load_workbook("data_table_test.xlsx")

    # Select the active worksheet
    work_sheet = data_work_book.active

    # Set the first cell of the first row to 'File Name'
    work_sheet.cell(row=1, column=1).value = 'File Name'

    # Add keywords to the first row
    for i, keyword in enumerate(keywords):
        work_sheet.cell(row=1, column=i + 2).value = keyword

    return data_work_book, work_sheet


def extract_text_from_pdf(pdf_file_path):
    parsedPDF = parser.from_file(pdf_file_path)
    return parsedPDF['content']


def create_text_file(pdf_file_path, extracted_text):
    #create the name of the new text file
    text_file_name = os.path.splitext(pdf_file_path)[0] + '.txt'

    with open(text_file_name, 'w', encoding='utf-8') as text_file:
        text_file.write(extracted_text)


def add_data_to_excel_table(worksheet, next_empty_row_number, pdf_file_name, keyword_counts_list):
    #put the PDF file name
    worksheet.cell(row=next_empty_row_number, column=1).value = pdf_file_name
    for i, count in enumerate(keyword_counts_list):
        worksheet.cell(row=next_empty_row_number, column=i + 2).value = count

#add the 'topic' column to the data table
def add_topic_to_table(worksheet, topic):
    # Create a new column titled 'Topic'
    worksheet.cell(row=1, column=len(keywords) + 2).value = 'Topic'
    
    # Add each topic to the new column
    for i, topic in enumerate(topic):
        worksheet.cell(row=i + 2, column=len(keywords) + 2).value = topic
        
    return worksheet


def process_folder(main_folder, worksheet):

    # Find the next empty row in the worksheet - we'll put the final data there
    next_empty_row_number = worksheet.max_row + 1

    # loop through each folder in the main folder
    for sub_folder_name in os.listdir(main_folder):
        sub_folder_path = os.path.join(main_folder, sub_folder_name)
        # check if folderpath is a folder (not a file)
        if os.path.isdir(sub_folder_path):
            print('Extracting text from PDFs in folder:', sub_folder_name)
            # loop through each PDF file in the folder
            for pdf_file_name in os.listdir(sub_folder_path):
                if pdf_file_name.endswith('.pdf'):
                    pdf_file_path = os.path.join(sub_folder_path, pdf_file_name)
                    # extract text from the PDF file
                    extracted_text = extract_text_from_pdf(pdf_file_path)
                    print('Text extracted from file:', pdf_file_name)

                    # create a new text file with the same name as the PDF file
                    create_text_file(pdf_file_path, extracted_text)

                    # Count keyword occurrences in the extracted text
                    keyword_counts_list = [extracted_text.count(keyword) for keyword in keywords]

                    # Add the file name and keyword counts to the Excel table
                    add_data_to_excel_table(worksheet, next_empty_row_number, pdf_file_name, keyword_counts_list)
                    next_empty_row_number += 1

    return worksheet


# set the path to your main folder containing subfolders with PDF files
main_folder = 'C:/Users/lidor/Documents/Lidor/Lidor Study/Project Portal/data'

# load the Excel table
workbook, worksheet = create_excel_table()

#Process the main folder and add data to the Excel table
worksheet_with_keywords_counting = process_folder(main_folder, worksheet)

final_table_with_topics = add_topic_to_table(worksheet_with_keywords_counting, topic)

#Save the workbook
workbook.save('data_table.xlsx')
