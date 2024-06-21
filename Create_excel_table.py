from openpyxl import Workbook, load_workbook

work_book = load_workbook("C:/Users/lidor/Documents/Lidor/Lidor Study/Project Portal/data_table1.xlsx")
work_sheet = work_book.active
keywords = ['משחק', 'שחקן' , 'שחקנים' , 'קרב']
work_sheet['A2'].value = "test"
work_sheet.cell(row=1, column=2).value = keywords

work_book.save("data_table1.xlsx")