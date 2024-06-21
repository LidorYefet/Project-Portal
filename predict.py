import os
from openpyxl import load_workbook
from sklearn.preprocessing import LabelEncoder
import joblib
from tika import parser
import pandas as pd
import numpy as np  # you need to import numpy

# Create a new Excel workbook
data_work_book = load_workbook("data_table.xlsx")

# Select the active worksheet
work_sheet = data_work_book.active

# Load the decision tree model
model = joblib.load("projectPortal_tree_model.joblib")

# Load the label encoder
label_encoder = joblib.load("label_encoder.joblib")

keywords = ['משחק', 'שחקן' , 'שחקנים' , 'קרב',  'ניצחון',  'הפסד',  'Unity', 'גיימינג' , 'עלילה', 'פאזל', 'פאזלים' , 'עלילתי' ,
'חייזר' , 'שחקני' ,'שלב' , 'יוניטי', 'מסתורי', 'ניקוד', 'אסטרטגיה', 'יריב',  'תחרות', 'משימה',  'שלב', 'אויב' 'הרפתקאות',
'הרפתקה', 'מכשולים',  'להילחם', 'פסילה', 'חייל','משפטיים' , 'משפטיות' , 'משפטי', 'משפטית', 'עורך' , 'עורכי' , 'דין' , 'משפט' ,'זכות',
'זכויות' , 'תביעה' , 'חוקים', 'חקיקה' , 'נחקק' , 'חוק' , 'סעיף' , 'פרקליט', 'עדות' , 'עדויות'  'ראיות' , 'דיון' , 'דיונים' , 'תיק' ,
'מרצה' ,'מרצים','אקדמיה', 'מחלקה' , 'ועדה' , 'הרצאות', 'מצגות', 'מצגת', 'בתי ספר' , 'בתי הספר', 'בית הספר' , 'בית ספר' , 'לימודים' , 'תלמיד' ,
'קורס', 'מקצוע','מורים' ,'מורה', 'הוראה', 'שיעור', 'חינוך', 'כיתה', 'כיתתי', 'ילד' , 'ילדים', 'זיווג', 'הורות' , 'משפחה' , 'זוג' ,'לידה' , 'שידוך',
'שידוכים', 'זוגיות', 'בן זוג', 'בת זוג', 'היכרות', 'תינוק' , 'תור'  , 'רופא' , 'מטופל' , 'מטפל' , 'טיפול', 'רפואה' , 'רפואי' , 'מדד' , 'לחץ דם' , 'דופק' ,
'סטורציה' , 'נשימה' , 'נשימות', 'אבחון' , 'הפרעה' , 'הפרעות', 'אירידיולוגיה','מוקדים' , 'טלפוניים',  'טלפון', 'מייל' , 'דיווח', 'לדווח' , 'מוקד' , 'פניה' ,
'פניות' , 'מדווח' , 'להגיב' , 'מסמכים', 'טקסטים' , 'טקסט' , 'טקסטואליות', 'טקסטואלית', 'מסמך', 'תרגום', 'ביטוי', 'חילוץ', 'ביטויים', 'ארכיון', 'ארכיונים', 'קריאה',
'כבדי שמיעה', 'כבד השמיעה' , 'לקות',  'שפת הסימנים', 'חירש' , 'בעלי מוגבלות' , 'לקויות' , 'עיוורים' , 'אוטיזם' , 'אוטיסט' , 'נויורולוגי' , 'נחייה', 'אילם', 'אילמים', 
'קהל', 'אנשים', 'שאלונים' , 'שאלון', 'סקר', 'התנדבות' , 'אוכלוסייה' , 'מתנדב' , 'חברתי' , 'תוכן', 'תכנים' , 'שיתופי',
'רכישה', 'חנות' , 'מפעל' , 'יצרן' , 'יצרנים', 'מחיר' , 'עסק' ,'מסחר' , 'משלוח' , 'חשבונית' , 'חשבוניות' , 'יד שניה', 'מכירה', 'מוצר', 'מפרסם', 'עלות', 'תשלום',
'פריט', 'שוק', 'סופרמרקט', 'קופה', 'קופאי', 'ספק', 'חנויות', 'הזמנה', 'עובד' , 'כוח אדם' , 'מסגרת' , 'הכשרה' ,'מוזיקלי', 'מוזיקה' , 'קול' , 'אפקט', 'סאונד' , 'הופעה' ,
'תייר' ,'מסורת' , 'תרבות' , 'אטרקציות' , 'אטרקציה' , 'מסעדה'  , 'מסעדות' , 'מוזיאון' , 'סרט','Movie', 'אמן' , 'אמנים' ,'כרטיס', 'ספריה', 'מוזמנים', 'אירוע', 'סאונד', 
'תופים', 'נגינה','מזון', 'ירקות', 'פירות', 'בשר', 'חלב', 'מתכונים', 'מתכון',
'בריאות', 'כושר', 'אימון', 'מתאמן', 'תזונה', 'מאמנים', 'עישון',  'מעשן', 'גמילה', 'מעשנים', 'סיגריה','כנס', 'סיגריות',
'מתמטי' , 'פונקציה' , 'פונקציות' , 'אלגברה' , 'מטריצה' , 'דטרמיננטה' , 'גרף', 'גרפים', 'קודקוד',
'דירה' , 'דייר' , 'השכרה', 'השכרת', 'דירות' , 'שוכר', 'משכיר', 'חשמל','ארנונה', 'משכנתא', 'הוצאות', 'הכנסות', 'חסכונית', 'חשבון', 'חשבונות', 'דוח',  
'עיצוב' , 'לעצב' , 'יעצב' , 'תמונה' , 'תמונות' , 'פילטר' , 'filter' , 
'נהג' , 'צמיג' , 'רכב', 'מוסך' , 'נסיעות' , 'יעד' , 'ניווט' , 'waze' , 'לנסוע' ,'צמח','כביש' , 'דרך'  , 'דרכים' , 'הגה' , 'נהיגה' , 'נתיב', 'נוסע', 'נסיעה', 'מפה', 'מצפן', 'מיקום',
'מרחק', 'מסלול', 'טיול' , 'הליכה' , 'טיול' , 'כלב' , 'סביבה',
'סכנות' ,'מסוכן', 'סכנה' , 'נזק' , 'שריפה', 'שריפות', 'להתריע' , 'התרעה', 'חירום', 'הצלה', 'לכוד', 'לכודים',  'מצוקה', 'בריונות' , 'אלימות' , 'אלימה' , 'אלים' , 'התעללות' ,
'קוד', 'שגיאה', 'מתכנת',  'stack overflow', 'אובייקט', 'תחזוקה', 'שרת', 'חומרה']   

def extract_text_from_pdf(pdf_file_path):
    parsedPDF = parser.from_file(pdf_file_path)
    return parsedPDF['content']

def create_text_file(pdf_file_path, extracted_text):
    #create the name of the new text file
    text_file_name = os.path.splitext(pdf_file_path)[0] + '.txt'

    with open(text_file_name, 'w', encoding='utf-8') as text_file:
        text_file.write(extracted_text)

def add_data_to_excel_table(worksheet, next_empty_row_number, pdf_file_name, keyword_counts_list):
    worksheet.cell(row=next_empty_row_number, column=1).value = pdf_file_name
    for i, count in enumerate(keyword_counts_list):
        worksheet.cell(row=next_empty_row_number, column=i + 2).value = count

def classification(keyword_counts_list, model, label_encoder, worksheet, row_number):
    # Prepare the data for prediction
    X_pred = np.array(keyword_counts_list).reshape(1, -1)  # reshape your list

    # Use the model to predict the topic
    prediction = model.predict(X_pred)

    # Decode the prediction to get the original label
    predicted_topic = label_encoder.inverse_transform(prediction)[0]

    # Add the topic to the Excel table
    worksheet.cell(row=row_number, column=len(keywords) + 2).value = predicted_topic  # assuming topic column is after all keywords

    return worksheet


new_pdf_file_path = "newFiles/VioREACH.pdf"

new_pdf_file_name = os.path.basename(new_pdf_file_path)  # get the file name only, no full path

# extract text from the PDF file
new_extracted_text = extract_text_from_pdf(new_pdf_file_path)

# create a new text file with the same name as the PDF file
create_text_file(new_pdf_file_path, new_extracted_text)

# Count keyword occurrences in the extracted text
keyword_counts_list = [new_extracted_text.count(keyword) for keyword in keywords]

# Add the file name and keyword counts to the Excel table
add_data_to_excel_table(work_sheet, (work_sheet.max_row + 1) , new_pdf_file_name, keyword_counts_list)

# Classify the new document and add the topic to the Excel table
print()
work_sheet = classification(keyword_counts_list, model, label_encoder, work_sheet, (work_sheet.max_row))

# Save the workbook
data_work_book.save('data_table.xlsx')
